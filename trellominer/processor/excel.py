"""
Manipulate Excel workbooks.
"""
from datetime import datetime
import os

from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill

from trellominer.api import trello
from trellominer.config import yaml


class Excel(object):

    def __init__(self):
        self.config = yaml.read(os.getenv("TRELLO_CONFIG", default=os.path.join(os.path.expanduser('~'), ".trellominer.yaml")))
        self.output_file = os.getenv("TRELLO_OUTPUT_FILE", default=self.config['api']['output_file_name'])
        self.filename = os.path.join(os.path.expanduser('~'), "{0} {1}.xlsx".format(self.output_file, datetime.now().strftime("%Y-%m-%d")))

    def filename(self):
        return self.filename

    def process_boards(self, boards):
        """Creates a series of worksheets in a workbook based on Trello board names."""
        highlight = NamedStyle(name='highlight')
        highlight.font = Font(bold=True, size=20)

        tabletop = NamedStyle(name='tabletop')
        tabletop.font = Font(bold=True, color='FFFFFF')
        bd = Side(style='thin', color="000000", border_style='thin')
        tabletop.border = Border(left=bd, right=bd, top=bd, bottom=bd)
        tabletop.fill = PatternFill("solid", bgColor="333333")

        wb = Workbook()
        wb.add_named_style(highlight)

        lookup = trello.Trello()

        current_row = 6

        # Dump the default worksheet from the document.
        # TODO: (PS) Is there a better way to handle this?
        for sheet in wb:
            if "Sheet" in sheet.title:
                wb.remove_sheet(sheet)

        for board in boards:
            if "Projects" in board['name']:
                ws = wb.create_sheet(title="{0}".format(board['name']), index=0)
                ws.sheet_properties.tabColor = "0000FF"
            elif "Break Fix" in board['name']:
                ws = wb.create_sheet(title="{0}".format(board['name']), index=1)
                ws.sheet_properties.tabColor = "FF0000"
            elif "Change Control" in board['name']:
                ws = wb.create_sheet(title="{0}".format(board['name']), index=2)
                ws.sheet_properties.tabColor = "228B22"
            else:
                ws = wb.create_sheet(title="{0}".format(board['name'][0:30]), index=None)
            ws['A1'].style = 'highlight'
            ws['A1'] = "{0}".format(board['name'])

            ws['A2'] = ""   # was going to contain board descriptions. Trello have deprecated these, just not from the API
            ws['A3'] = "{0}".format(board['url'])
            ws['A3'].style = 'Hyperlink'
            ws['A4'] = ""

            headings = ["Name", "Description", "Status", "Due Date", "Complete", "Perc", "Members"]
            ws.append(headings)
            header_row = ws[5]
            for cell in header_row:
                cell.style = tabletop

            cards = lookup.cards(board['shortLink'])

            # Apply some default column widths to each worksheet
            ws.column_dimensions["A"].width = 40
            ws.column_dimensions["B"].width = 100
            ws.column_dimensions["C"].width = 10
            ws.column_dimensions["D"].width = 22
            ws.column_dimensions["G"].width = 45

            for card in cards:
                # TODO: Pretty slow to iterate like this. Improve.
                listname = lookup.lists(card['idList'])

                member_list = ""
                for member in card['members']:
                    member_list += "{0},".format(member['fullName'])
                member_list.replace(',', ', ')

                ws["A{0}".format(current_row)] = card['name']
                ws["A{0}".format(current_row)].style = 'Output'
                ws["B{0}".format(current_row)] = card['desc']
                ws["C{0}".format(current_row)] = listname['name']
                if 'Conceptual' in listname['name']:
                    ws["C{0}".format(current_row)].style = 'Accent5'
                elif 'Backlog' in listname['name']:
                    ws["C{0}".format(current_row)].style = 'Accent4'
                elif 'In Progress' in listname['name']:
                    ws["C{0}".format(current_row)].style = 'Accent1'
                elif 'Impeded' in listname['name']:
                    ws["C{0}".format(current_row)].style = 'Accent6'
                elif 'Completed' in listname['name']:
                    ws["C{0}".format(current_row)].style = 'Accent3'
                elif 'Stopped' in listname['name']:
                    ws["C{0}".format(current_row)].style = 'Accent2'
                elif 'Planned' in listname['name']:
                    ws["C{0}".format(current_row)].style = 'Accent4'
                elif 'Successful' in listname['name']:
                    ws["C{0}".format(current_row)].style = 'Good'
                elif 'Failed' in listname['name']:
                    ws["C{0}".format(current_row)].style = 'Bad'
                elif 'Cancelled' in listname['name']:
                    ws["C{0}".format(current_row)].style = 'Neutral'
                else:
                    ws["C{0}".format(current_row)] = listname['name']

                ws["D{0}".format(current_row)] = card['due']
                ws["E{0}".format(current_row)] = card['dueComplete']
                # ws["F{0}".format(current_row)] = card['closed']
                tasks = 0
                complete = 0
                checklists = lookup.checklists(card['shortLink'])
                for checklist in checklists:
                    for cl in checklist['checkItems']:
                        tasks += 1
                        if cl['state'] == 'complete':
                            complete += 1
                if tasks > 0:
                    perc = 100 * complete / tasks
                else:
                    perc = 0
                ws["F{0}".format(current_row)] = "{0}%".format(perc)
                if perc < 25:
                    ws["F{0}".format(current_row)].style = 'Bad'
                elif perc < 50:
                    ws["F{0}".format(current_row)].style = 'Neutral'
                else:
                    ws["F{0}".format(current_row)].style = 'Good'
                ws["G{0}".format(current_row)] = member_list[:-1]
                current_row += 1

            current_row = 6

        wb.save(self.filename)
